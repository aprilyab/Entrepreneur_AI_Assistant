"""
Export utilities for investor-ready business plan documents.
"""
from __future__ import annotations

import html
import re
from datetime import datetime
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.graphics.shapes import Circle, Drawing, Rect, String
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


INK = colors.HexColor("#14213D")
MUTED = colors.HexColor("#667085")
TEAL = colors.HexColor("#0F766E")
TEAL_LIGHT = colors.HexColor("#E8F5F2")
CORAL = colors.HexColor("#F97360")
PAPER = colors.HexColor("#F7F5F0")
LINE = colors.HexColor("#D9E1E7")
WHITE = colors.white


def _register_fonts() -> tuple[str, str]:
    """Use a modern system font when available, with safe built-in fallbacks."""
    regular = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    bold = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    try:
        pdfmetrics.registerFont(TTFont("ExportSans", regular))
        pdfmetrics.registerFont(TTFont("ExportSans-Bold", bold))
        return "ExportSans", "ExportSans-Bold"
    except Exception:
        return "Helvetica", "Helvetica-Bold"


FONT, FONT_BOLD = _register_fonts()


def _styles() -> dict[str, ParagraphStyle]:
    sample = getSampleStyleSheet()
    return {
        "cover_label": ParagraphStyle(
            "CoverLabel",
            parent=sample["Normal"],
            fontName=FONT_BOLD,
            fontSize=9,
            leading=12,
            textColor=TEAL,
            tracking=1.3,
            alignment=TA_CENTER,
            spaceAfter=16,
        ),
        "cover_title": ParagraphStyle(
            "CoverTitle",
            parent=sample["Title"],
            fontName=FONT_BOLD,
            fontSize=30,
            leading=36,
            textColor=INK,
            alignment=TA_CENTER,
            spaceAfter=12,
        ),
        "cover_subtitle": ParagraphStyle(
            "CoverSubtitle",
            parent=sample["Normal"],
            fontName=FONT,
            fontSize=12,
            leading=18,
            textColor=MUTED,
            alignment=TA_CENTER,
        ),
        "h1": ParagraphStyle(
            "H1",
            parent=sample["Heading1"],
            fontName=FONT_BOLD,
            fontSize=21,
            leading=26,
            textColor=INK,
            spaceBefore=4,
            spaceAfter=12,
            keepWithNext=True,
        ),
        "h2": ParagraphStyle(
            "H2",
            parent=sample["Heading2"],
            fontName=FONT_BOLD,
            fontSize=14,
            leading=18,
            textColor=TEAL,
            spaceBefore=14,
            spaceAfter=7,
            keepWithNext=True,
        ),
        "h3": ParagraphStyle(
            "H3",
            parent=sample["Heading3"],
            fontName=FONT_BOLD,
            fontSize=11,
            leading=15,
            textColor=INK,
            spaceBefore=10,
            spaceAfter=5,
            keepWithNext=True,
        ),
        "body": ParagraphStyle(
            "Body",
            parent=sample["BodyText"],
            fontName=FONT,
            fontSize=9.4,
            leading=14,
            textColor=INK,
            spaceAfter=7,
        ),
        "bullet": ParagraphStyle(
            "Bullet",
            parent=sample["BodyText"],
            fontName=FONT,
            fontSize=9.2,
            leading=13.5,
            textColor=INK,
            leftIndent=15,
            firstLineIndent=-9,
            bulletIndent=4,
            spaceAfter=4,
        ),
        "small": ParagraphStyle(
            "Small",
            parent=sample["BodyText"],
            fontName=FONT,
            fontSize=8,
            leading=11,
            textColor=MUTED,
        ),
        "toc": ParagraphStyle(
            "Contents",
            parent=sample["BodyText"],
            fontName=FONT,
            fontSize=10.5,
            leading=16,
            textColor=INK,
            leftIndent=9,
        ),
        "metric": ParagraphStyle(
            "Metric",
            parent=sample["Normal"],
            fontName=FONT_BOLD,
            fontSize=22,
            leading=25,
            textColor=TEAL,
            alignment=TA_CENTER,
        ),
        "metric_label": ParagraphStyle(
            "MetricLabel",
            parent=sample["Normal"],
            fontName=FONT,
            fontSize=7.5,
            leading=10,
            textColor=MUTED,
            alignment=TA_CENTER,
        ),
    }


def _inline_markdown(value: str) -> str:
    value = html.escape(value.strip())
    value = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", value)
    value = re.sub(r"(?<!\*)\*([^*]+?)\*(?!\*)", r"<i>\1</i>", value)
    value = re.sub(r"`([^`]+)`", r"<font name='Courier'>\1</font>", value)
    return value


def _plain(value: str) -> str:
    value = re.sub(r"\*\*(.+?)\*\*", r"\1", value)
    value = re.sub(r"(?<!\*)\*([^*]+?)\*(?!\*)", r"\1", value)
    value = re.sub(r"`([^`]+)`", r"\1", value)
    return value.strip()


def _extract(pattern: str, text: str, default: str = "—") -> str:
    match = re.search(pattern, text or "", re.IGNORECASE)
    return _plain(match.group(1)).strip(" .") if match else default


def _label_line(text: str, label: str) -> str:
    for raw in (text or "").splitlines():
        clean = _plain(raw)
        position = clean.lower().find(label.lower())
        if position < 0:
            continue
        separator = clean.find(":", position + len(label))
        return clean[separator + 1 :].strip(" -*") if separator >= 0 else clean
    return ""


def _currency_label(text: str, label: str, *, last: bool = False, default: str = "—") -> str:
    values = re.findall(
        r"\$[\d,.]+(?:\s*(?:billion|million|B|M))?",
        _label_line(text, label),
        re.IGNORECASE,
    )
    if not values:
        return default
    value = values[-1] if last else values[0]
    value = re.sub(r"\s+billion\b", "B", value, flags=re.IGNORECASE)
    return re.sub(r"\s+million\b", "M", value, flags=re.IGNORECASE)


def _section(markdown: str, *needles: str) -> str:
    sections: dict[str, list[str]] = {"overview": []}
    current = "overview"
    for raw in (markdown or "").splitlines():
        match = re.match(r"^#{1,4}\s+(?:\d+[.)]\s*)?(.+)$", raw.strip())
        if match:
            current = _plain(match.group(1)).lower()
            sections.setdefault(current, [])
        else:
            sections.setdefault(current, []).append(raw)
    for key, value in sections.items():
        if any(needle.lower() in key for needle in needles):
            return "\n".join(value)
    return ""


def _bullets(text: str, limit: int = 4) -> list[str]:
    result: list[str] = []
    for raw in (text or "").splitlines():
        match = re.match(r"^\s*(?:[-*]|\d+[.)])\s+(.+)$", raw)
        if match:
            result.append(_plain(match.group(1)))
        if len(result) == limit:
            break
    return result


def _labeled_bullets(text: str, label: str, limit: int = 3) -> list[str]:
    lines = (text or "").splitlines()
    start = next(
        (
            index
            for index, line in enumerate(lines)
            if re.sub(r"^[-*]\s*", "", _plain(line)).lower().startswith(f"{label.lower()}:")
        ),
        -1,
    )
    if start < 0:
        return []
    result: list[str] = []
    for raw in lines[start + 1 :]:
        if re.match(r"^\s*[-*]\s+\*\*[^*]+:\*\*\s*$", raw):
            break
        match = re.match(r"^\s*[-*]\s+(.+)$", raw)
        if match:
            result.append(_plain(match.group(1)))
        if len(result) == limit:
            break
    return result


def _visual_title(text: str, styles: dict[str, ParagraphStyle]):
    return Paragraph(text, styles["h3"])


def _wrap_words(value: str, width: int = 24, max_lines: int = 3) -> list[str]:
    words = value.split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) <= width:
            current = candidate
            continue
        if current:
            lines.append(current)
        current = word
        if len(lines) == max_lines:
            break
    if current and len(lines) < max_lines:
        lines.append(current)
    consumed = " ".join(lines)
    if len(consumed) < len(value.strip()) and lines:
        lines[-1] = lines[-1].rstrip(" ,.;:") + "…"
    return lines


def _market_figure(text: str, width: float, styles: dict[str, ParagraphStyle]) -> list:
    values = [
        ("TAM", _currency_label(text, "TAM (Total Addressable Market)", last=True), colors.HexColor("#CFFAFE")),
        ("SAM", _currency_label(text, "SAM (Serviceable Available Market)"), colors.HexColor("#C7D2FE")),
        ("SOM", _currency_label(text, "SOM (Serviceable Obtainable Market)"), INK),
    ]
    drawing = Drawing(width, 150)
    centers = [(105, 76, 64), (255, 76, 50), (380, 76, 37)]
    for (label, value, color), (x, y, radius) in zip(values, centers):
        drawing.add(Circle(x, y, radius, fillColor=color, strokeColor=colors.white, strokeWidth=3))
        text_color = colors.white if label == "SOM" else INK
        drawing.add(String(x, y + 7, label, textAnchor="middle", fontName=FONT_BOLD, fontSize=8, fillColor=text_color))
        drawing.add(String(x, y - 12, value, textAnchor="middle", fontName=FONT_BOLD, fontSize=14, fillColor=text_color))
    growth = re.search(r"(\d+(?:\.\d+)?%\s*(?:to|–|-)\s*\d+(?:\.\d+)?%)", text or "")
    drawing.add(String(width - 2, 80, growth.group(1) if growth else "Growth not provided", textAnchor="end", fontName=FONT_BOLD, fontSize=11, fillColor=TEAL))
    drawing.add(String(width - 2, 64, "ESTIMATED CAGR", textAnchor="end", fontName=FONT, fontSize=7, fillColor=MUTED))
    return [_visual_title("Market sizing snapshot", styles), drawing, Spacer(1, 8)]


def _model_figure(text: str, width: float, styles: dict[str, ParagraphStyle]) -> list:
    drawing = Drawing(width, 142)
    groups = [
        ("CUSTOMERS", _bullets(_section(text, "customer segments"), 1) or _labeled_bullets(text, "Customer Segments", 1), colors.HexColor("#ECFEFF")),
        ("CHANNELS", _bullets(_section(text, "channels"), 1) or _labeled_bullets(text, "Channels", 1), colors.HexColor("#EEF2FF")),
        ("REVENUE", _bullets(_section(text, "revenue streams"), 1) or _labeled_bullets(text, "Revenue Streams", 1), colors.HexColor("#ECFDF5")),
    ]
    box_width = (width - 48) / 3
    for index, (label, values, fill) in enumerate(groups):
        x = index * (box_width + 24)
        drawing.add(Rect(x, 20, box_width, 105, rx=8, ry=8, fillColor=fill, strokeColor=LINE))
        drawing.add(String(x + 12, 101, label, fontName=FONT_BOLD, fontSize=8, fillColor=TEAL))
        body = values[0] if values else "Not structured"
        for line_index, line in enumerate(_wrap_words(body)):
            drawing.add(
                String(
                    x + 12,
                    73 - line_index * 14,
                    line,
                    fontName=FONT_BOLD if line_index == 0 else FONT,
                    fontSize=8,
                    fillColor=INK,
                )
            )
        if index < 2:
            drawing.add(String(x + box_width + 6, 70, "→", fontName=FONT_BOLD, fontSize=17, fillColor=MUTED))
    return [_visual_title("Value-flow summary", styles), drawing, Spacer(1, 8)]


def _score_figure(text: str, width: float, styles: dict[str, ParagraphStyle]) -> list:
    rows: list[tuple[str, int, int]] = []
    for raw in (text or "").splitlines():
        if raw.strip().startswith("#"):
            continue
        clean = _plain(re.sub(r"^\s*[-*]\s+", "", raw))
        match = re.match(r"^([^:]+):\s*(\d+)\s*/\s*(\d+)", clean)
        if match:
            rows.append((match.group(1), int(match.group(2)), int(match.group(3))))
    rows = rows[:6]
    drawing = Drawing(width, max(85, len(rows) * 30 + 18))
    y = drawing.height - 22
    for label, value, maximum in rows:
        drawing.add(String(0, y + 2, label[:24], fontName=FONT_BOLD, fontSize=8, fillColor=INK))
        drawing.add(Rect(132, y, width - 178, 9, rx=4, ry=4, fillColor=colors.HexColor("#E8EEF2"), strokeColor=None))
        drawing.add(Rect(132, y, (width - 178) * value / maximum, 9, rx=4, ry=4, fillColor=TEAL, strokeColor=None))
        drawing.add(String(width - 4, y + 1, f"{value}/{maximum}", textAnchor="end", fontName=FONT_BOLD, fontSize=8, fillColor=INK))
        y -= 30
    return [_visual_title("Decision score anatomy", styles), drawing, Spacer(1, 8)] if rows else []


def _risk_figure(text: str, width: float, styles: dict[str, ParagraphStyle]) -> list:
    names: list[str] = []
    probabilities: list[str] = []
    impacts: list[str] = []
    group_impact = "Medium"
    for raw in (text or "").splitlines():
        if re.match(r"^#{1,4}\s+(?:Critical|High) Risks", raw.strip(), re.IGNORECASE):
            group_impact = "High"
        elif re.match(r"^#{1,4}\s+Medium Risks", raw.strip(), re.IGNORECASE):
            group_impact = "Medium"
        clean = _plain(re.sub(r"^\s*[-*]\s+", "", raw))
        if clean.lower().startswith("risk:"):
            names.append(clean.split(":", 1)[1].strip())
            impacts.append(group_impact)
        elif clean.lower().startswith("probability:"):
            probabilities.append(clean.split(":", 1)[1].strip())
        elif clean.lower().startswith("impact:"):
            value = clean.split(":", 1)[1].strip()
            if impacts and re.match(r"^(?:high|critical)\b", value, re.I):
                impacts[-1] = "High"
            elif impacts and re.match(r"^low\b", value, re.I):
                impacts[-1] = "Low"
    drawing = Drawing(width, 170)
    cell_w, cell_h = 105, 43
    origin_x, origin_y = 95, 25
    palette = [
        [colors.HexColor("#ECFDF5"), colors.HexColor("#FEF3C7"), colors.HexColor("#FEE2E2")],
        [colors.HexColor("#ECFDF5"), colors.HexColor("#FEF3C7"), colors.HexColor("#FECACA")],
        [colors.HexColor("#F8FAFC"), colors.HexColor("#FEF3C7"), colors.HexColor("#FEE2E2")],
    ]
    for row in range(3):
        for column in range(3):
            drawing.add(Rect(origin_x + column * cell_w, origin_y + row * cell_h, cell_w, cell_h, fillColor=palette[row][column], strokeColor=colors.white))
    drawing.add(String(origin_x + 1.5 * cell_w, 5, "IMPACT  →", textAnchor="middle", fontName=FONT_BOLD, fontSize=7, fillColor=MUTED))
    drawing.add(String(17, origin_y + 1.5 * cell_h, "PROBABILITY", fontName=FONT_BOLD, fontSize=7, fillColor=MUTED))
    levels = {"low": 0, "medium": 1, "high": 2}
    for index, name in enumerate(names[:6]):
        probability = probabilities[index] if index < len(probabilities) else "Medium"
        impact = impacts[index] if index < len(impacts) else "Medium"
        column = levels.get(impact.lower(), 1)
        row = levels.get(probability.lower(), 1)
        x = origin_x + column * cell_w + 18 + (index % 2) * 28
        y = origin_y + row * cell_h + 18
        drawing.add(Circle(x, y, 11, fillColor=INK, strokeColor=colors.white, strokeWidth=2))
        drawing.add(String(x, y - 3, str(index + 1), textAnchor="middle", fontName=FONT_BOLD, fontSize=8, fillColor=colors.white))
    return [_visual_title("Risk exposure map", styles), drawing, Spacer(1, 8)]


def _financial_figure(text: str, width: float, styles: dict[str, ParagraphStyle]) -> list:
    patterns = [
        r"Year 1(?:\s+Projections)?[\s\S]{0,260}?Total Revenue:[^$\n]*\$([0-9,.]+)",
        r"Year 2 Total Revenue:[^$\n]*\$([0-9,.]+)",
        r"Year 3[\s\S]{0,120}?Total Revenue:[^$\n]*\$([0-9,.]+)",
    ]
    values: list[float] = []
    labels: list[str] = []
    for pattern in patterns:
        match = re.search(pattern, text or "", re.IGNORECASE)
        raw = match.group(1) if match else "0"
        values.append(float(raw.replace(",", "")))
        labels.append(f"${raw}" if match else "Not provided")
    maximum = max(max(values or [0]), 1)
    drawing = Drawing(width, 175)
    colors_by_year = [colors.HexColor("#4F46E5"), colors.HexColor("#06B6D4"), colors.HexColor("#10B981")]
    for index, value in enumerate(values):
        bar_height = max(8, 105 * value / maximum)
        x = 55 + index * 145
        drawing.add(Rect(x, 30, 70, bar_height, rx=5, ry=5, fillColor=colors_by_year[index], strokeColor=None))
        drawing.add(String(x + 35, 17, f"YEAR {index + 1}", textAnchor="middle", fontName=FONT_BOLD, fontSize=7, fillColor=MUTED))
        drawing.add(String(x + 35, 39 + bar_height, labels[index], textAnchor="middle", fontName=FONT_BOLD, fontSize=9, fillColor=INK))
    drawing.add(String(width - 6, 149, "PROJECTED REVENUE", textAnchor="end", fontName=FONT_BOLD, fontSize=7, fillColor=TEAL))
    return [_visual_title("Three-year revenue trajectory", styles), drawing, Spacer(1, 8)]


def _markdown_table(lines: list[str], styles: dict[str, ParagraphStyle], width: float):
    rows: list[list[str]] = []
    for line in lines:
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in cells):
            continue
        rows.append(cells)
    if not rows:
        return Spacer(1, 1)

    columns = max(len(row) for row in rows)
    normalized = [row + [""] * (columns - len(row)) for row in rows]
    data = [
        [Paragraph(_inline_markdown(cell), styles["small"]) for cell in row]
        for row in normalized
    ]
    table = Table(data, colWidths=[width / columns] * columns, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), INK),
                ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.45, LINE),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _markdown_flowables(
    content: str,
    styles: dict[str, ParagraphStyle],
    available_width: float,
    *,
    skip_first_h1: bool = False,
) -> list:
    """Turn the limited Markdown produced by the agents into ReportLab flowables."""
    output: list = []
    lines = (content or "").replace("\r\n", "\n").split("\n")
    paragraph_buffer: list[str] = []
    first_h1_seen = False

    def flush_paragraph():
        if paragraph_buffer:
            joined = " ".join(part.strip() for part in paragraph_buffer if part.strip())
            if joined:
                output.append(Paragraph(_inline_markdown(joined), styles["body"]))
            paragraph_buffer.clear()

    index = 0
    while index < len(lines):
        raw = lines[index].rstrip()
        stripped = raw.strip()

        if stripped.startswith("|") and "|" in stripped[1:]:
            flush_paragraph()
            table_lines = []
            while index < len(lines) and lines[index].strip().startswith("|"):
                table_lines.append(lines[index].strip())
                index += 1
            output.extend([_markdown_table(table_lines, styles, available_width), Spacer(1, 8)])
            continue

        heading = re.match(r"^(#{1,4})\s+(.+)$", stripped)
        if heading:
            flush_paragraph()
            level = len(heading.group(1))
            title = _plain(heading.group(2))
            if level == 1 and skip_first_h1 and not first_h1_seen:
                first_h1_seen = True
            else:
                first_h1_seen = first_h1_seen or level == 1
                style = styles["h1"] if level == 1 else styles["h2"] if level == 2 else styles["h3"]
                output.append(Paragraph(_inline_markdown(title), style))
            index += 1
            continue

        bullet = re.match(r"^(\s*)[-*]\s+(.+)$", raw)
        numbered = re.match(r"^(\s*)(\d+)[.)]\s+(.+)$", raw)
        if bullet or numbered:
            flush_paragraph()
            if bullet:
                indent, text, marker = len(bullet.group(1)), bullet.group(2), "•"
            else:
                indent, text, marker = len(numbered.group(1)), numbered.group(3), f"{numbered.group(2)}."
            style = ParagraphStyle(
                f"Bullet{indent}",
                parent=styles["bullet"],
                leftIndent=15 + min(indent, 12),
                bulletIndent=4 + min(indent, 12),
            )
            output.append(Paragraph(_inline_markdown(text), style, bulletText=marker))
            index += 1
            continue

        if stripped == "---":
            flush_paragraph()
            output.append(Spacer(1, 5))
            index += 1
            continue

        if not stripped:
            flush_paragraph()
        else:
            paragraph_buffer.append(stripped)
        index += 1

    flush_paragraph()
    return output


def _page_header_footer(canvas, doc, title: str):
    canvas.saveState()
    page = canvas.getPageNumber()
    if page > 1:
        canvas.setStrokeColor(LINE)
        canvas.setLineWidth(0.5)
        canvas.line(doc.leftMargin, letter[1] - 0.53 * inch, letter[0] - doc.rightMargin, letter[1] - 0.53 * inch)
        canvas.setFont(FONT_BOLD, 7.5)
        canvas.setFillColor(INK)
        canvas.drawString(doc.leftMargin, letter[1] - 0.42 * inch, title[:70])
        canvas.setFont(FONT, 7.5)
        canvas.setFillColor(MUTED)
        canvas.drawRightString(letter[0] - doc.rightMargin, 0.42 * inch, f"Business Decision Report  •  {page}")
    canvas.restoreState()


def export_plan_as_pdf(plan_data: dict) -> bytes:
    """Export a structured business report without raw Markdown or duplicated sections."""
    buffer = BytesIO()
    title = (plan_data.get("title") or plan_data.get("idea") or "Business Plan").strip()
    full_plan = plan_data.get("full_plan") or ""
    validation = plan_data.get("validation_strategy") or ""
    financials = plan_data.get("financials") or ""
    score = plan_data.get("viability_score") or _extract(r"Viability Score:\s*(\d+/?100)", validation)
    verdict = _extract(r"Verdict:\s*([^\n#]+)", validation, "Decision pending")
    styles = _styles()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.68 * inch,
        leftMargin=0.68 * inch,
        topMargin=0.72 * inch,
        bottomMargin=0.68 * inch,
        title=title,
        author="Business Decision Report",
    )
    usable_width = letter[0] - doc.leftMargin - doc.rightMargin
    story: list = []

    # Cover
    story.extend(
        [
            Spacer(1, 1.15 * inch),
            Paragraph("BUSINESS DECISION REPORT", styles["cover_label"]),
            Paragraph(_inline_markdown(title), styles["cover_title"]),
            Spacer(1, 0.12 * inch),
            Table(
                [
                    [
                        Paragraph(str(score), styles["metric"]),
                        Paragraph(verdict.upper(), styles["metric"]),
                    ],
                    [
                        Paragraph("VIABILITY SCORE", styles["metric_label"]),
                        Paragraph("RECOMMENDATION", styles["metric_label"]),
                    ],
                ],
                colWidths=[2.15 * inch, 2.65 * inch],
                rowHeights=[0.48 * inch, 0.27 * inch],
                hAlign="CENTER",
                style=TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), TEAL_LIGHT),
                        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#B8DCD6")),
                        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B8DCD6")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 0), (-1, -1), 7),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ]
                ),
            ),
            Spacer(1, 0.45 * inch),
            Paragraph(
                "A structured review of the opportunity, business model, execution plan, financial assumptions, and material risks.",
                styles["cover_subtitle"],
            ),
            Spacer(1, 1.25 * inch),
            Paragraph(datetime.now().strftime("%B %Y"), styles["cover_subtitle"]),
            PageBreak(),
        ]
    )

    # Contents and reading guide
    story.extend(
        [
            Paragraph("How to read this report", styles["h1"]),
            Paragraph(
                "The core business plan comes first. Detailed agent analyses follow as appendices, so the report remains useful without repeating the same material in the main narrative.",
                styles["body"],
            ),
            Spacer(1, 8),
            Table(
                [
                    [Paragraph("<b>01</b>", styles["toc"]), Paragraph("Core business plan", styles["toc"])],
                    [Paragraph("<b>02</b>", styles["toc"]), Paragraph("Market research appendix", styles["toc"])],
                    [Paragraph("<b>03</b>", styles["toc"]), Paragraph("Business model appendix", styles["toc"])],
                    [Paragraph("<b>04</b>", styles["toc"]), Paragraph("Validation and decision appendix", styles["toc"])],
                    [Paragraph("<b>05</b>", styles["toc"]), Paragraph("Risk register appendix", styles["toc"])],
                    [Paragraph("<b>06</b>", styles["toc"]), Paragraph("Financial model appendix", styles["toc"])],
                ],
                colWidths=[0.55 * inch, usable_width - 0.55 * inch],
                style=TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), PAPER),
                        ("LINEBELOW", (0, 0), (-1, -2), 0.4, LINE),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
                    ]
                ),
            ),
            PageBreak(),
            Paragraph("Core business plan", styles["h1"]),
        ]
    )
    story.extend(_markdown_flowables(full_plan, styles, usable_width, skip_first_h1=True))

    appendices = [
        ("Appendix A", "Market research", plan_data.get("market_analysis")),
        ("Appendix B", "Business model detail", plan_data.get("business_model")),
        ("Appendix C", "Validation and decision", validation),
        ("Appendix D", "Risk register", plan_data.get("risks")),
        ("Appendix E", "Financial model", financials),
    ]
    for label, heading, content in appendices:
        if not content:
            continue
        story.extend(
            [
                PageBreak(),
                Paragraph(label.upper(), styles["cover_label"]),
                Paragraph(heading, styles["h1"]),
            ]
        )
        if label == "Appendix A":
            story.extend(_market_figure(content, usable_width, styles))
        elif label == "Appendix B":
            story.extend(_model_figure(content, usable_width, styles))
        elif label == "Appendix C":
            story.extend(_score_figure(content, usable_width, styles))
        elif label == "Appendix D":
            story.extend(_risk_figure(content, usable_width, styles))
        elif label == "Appendix E":
            story.extend(_financial_figure(content, usable_width, styles))
        story.extend(_markdown_flowables(content, styles, usable_width))

    doc.build(
        story,
        onFirstPage=lambda canvas, d: _page_header_footer(canvas, d, title),
        onLaterPages=lambda canvas, d: _page_header_footer(canvas, d, title),
    )
    buffer.seek(0)
    return buffer.getvalue()


def export_financials_as_excel(financials: str, metrics: dict | None = None) -> bytes:
    """Export typed financial assumptions without inventing empty monthly values."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    metrics = metrics or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Key Metrics"
    headers = ["Metric", "Value", "Display", "Unit", "Basis", "Confidence", "Data type"]
    header_fill = PatternFill("solid", fgColor="14213D")
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    excluded = {"currency", "use_of_funds", "assumptions", "ltv_cac_denominator"}
    row = 2
    for key, item in metrics.items():
        if key in excluded or not isinstance(item, dict) or "value" not in item:
            continue
        values = [
            key.replace("_", " ").title(),
            item.get("value"),
            item.get("display"),
            item.get("unit"),
            item.get("basis"),
            item.get("confidence"),
            "Assumption" if item.get("assumption", True) else "Founder/historical",
        ]
        for column, value in enumerate(values, 1):
            ws.cell(row=row, column=column, value=value)
        row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="No structured metrics were available for this legacy plan.")

    allocations = wb.create_sheet("Use of Funds")
    for column, header in enumerate(["Category", "Percentage", "Amount"], 1):
        cell = allocations.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for index, item in enumerate(metrics.get("use_of_funds") or [], 2):
        allocations.cell(index, 1, item.get("category"))
        allocations.cell(index, 2, item.get("percentage"))
        allocations.cell(index, 3, item.get("amount"))

    assumptions = wb.create_sheet("Assumptions")
    assumption_headers = ["Assumption", "Value", "Rationale", "Confidence", "Impact", "Validation method"]
    for column, header in enumerate(assumption_headers, 1):
        cell = assumptions.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for index, item in enumerate(metrics.get("assumptions") or [], 2):
        for column, key in enumerate(
            ["name", "value", "rationale", "confidence", "impact", "validation_method"],
            1,
        ):
            assumptions.cell(index, column, item.get(key))

    narrative = wb.create_sheet("Narrative")
    narrative.cell(row=1, column=1, value="Generated financial analysis")
    narrative.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    narrative.cell(row=1, column=1).fill = header_fill
    for index, line in enumerate((financials or "No financial narrative available.").splitlines(), 2):
        narrative.cell(row=index, column=1, value=_plain(line))

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions if sheet.max_column > 1 else None
        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            width = min(
                55,
                max(
                    12,
                    max(
                        len(str(sheet.cell(row_index, column).value or ""))
                        for row_index in range(1, sheet.max_row + 1)
                    )
                    + 2,
                ),
            )
            sheet.column_dimensions[letter].width = width
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
